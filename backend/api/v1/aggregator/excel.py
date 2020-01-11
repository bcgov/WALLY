from typing import List
import openpyxl
from openpyxl.writer.excel import save_virtual_workbook
from starlette.responses import Response
from geojson import FeatureCollection
from api.v1.aggregator.schema import LayerResponse


def xlsxExport(features: List[LayerResponse]):
    """
    packages features into an excel workbook.  Returns an HTTP response object that has the saved workbook
    ready to be returned to the client (e.g. the calling http handler can return this object directly)
    """

    wb = openpyxl.Workbook()
    ws = wb.active
    first_sheet = True

    for dataset in features:
        # avoid trying to process layers if they have no features.
        if not dataset.geojson:
            continue

        # create a list of fields for this dataset
        fields = []
        try:
            fields = [*dataset.geojson[0].properties]
        except:
            continue

        if not first_sheet:
            ws = wb.create_sheet(dataset.layer)
        else:
            ws.title = dataset.layer
            first_sheet = False

        ws.append(fields)

        features = dataset.geojson.features

        # add rows for every object in the collection, using the fields defined above.
        for f in features:
            props = f['properties']
            ws.append([
                props.get(x) for x in fields
            ])

    response = Response(
        content=save_virtual_workbook(wb),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=report.xlsx'})

    return response


def geojson_to_xlsx(fc_list: List[FeatureCollection]):
    """
    packages a list of FeatureCollections into an excel workbook.
    Each FeatureCollection will get its own sheet/tab in the workbook.

    Returns an HTTP response object that has the saved workbook
    ready to be returned to the client (e.g. the calling http handler can return this object directly)
    """

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    first_sheet = True

    for i, fc in enumerate(fc_list, start=1):
        # avoid trying to process layers if they have no features.
        if not fc.features:
            continue

        # create a list of fields for this dataset
        fields = []
        try:
            fields = [*fc.features[0].properties]
        except:
            continue

        sheet_title = fc.properties['name'] or fc.properties['layer'] or f"sheet {i}"

        if not first_sheet:
            worksheet = workbook.create_sheet(sheet_title)
        else:
            worksheet.title = sheet_title
            first_sheet = False

        worksheet.append(fields)

        # add rows for every object in the collection, using the fields defined above.
        for f in fc.features:
            props = f['properties']
            worksheet.append([
                props.get(x) for x in fields
            ])

    response = Response(
        content=save_virtual_workbook(workbook),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=report.xlsx'})

    return response
