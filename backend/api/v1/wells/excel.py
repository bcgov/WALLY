from typing import List, Dict
import logging
import datetime
import openpyxl
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from starlette.responses import Response
from geojson import FeatureCollection
from api.v1.aggregator.schema import LayerResponse

logger = logging.getLogger('well export')


def cross_section_xlsx_export(features: List[LayerResponse], coordinates: list, buffer: int):
    """
    packages features into an excel workbook.  Returns an HTTP response object that has the saved workbook
    ready to be returned to the client (e.g. the calling http handler can return this object directly)
    """

    workbook = openpyxl.Workbook()
    ds = workbook.active
    ds.title = "details"

    font_title = Font(size=20, bold=True, color='44546a')
    font_label = Font(bold=True)
    border_bottom = Border(bottom=Side(border_style="thick", color='4472c4'))

    # styling config
    ds['A1'].font = font_title
    ds.column_dimensions['A'].width = 20
    ds['A1'].border = border_bottom
    ds['B1'].border = border_bottom
    ds['C1'].border = border_bottom

    ds['A2'].font = font_label
    ds['A3'].font = font_label
    ds['A4'].font = font_label
    ds['A5'].font = font_label

    ds['A1'] = 'Cross section'
    ds['A2'] = 'Date generated:'
    ds['A3'] = 'A point coordinates:'
    ds['A4'] = 'B point coordinates:'
    ds['A5'] = 'Buffer radius (m):'

    cur_date = datetime.datetime.now().strftime("%X-%Y-%m-%d")

    ds['B2'] = cur_date
    ds['B3'] = str(coordinates[0][1]) + ', ' + str(coordinates[0][0])
    ds['B4'] = str(coordinates[1][1]) + ', ' + str(coordinates[1][0])
    ds['B5'] = buffer

    # create data sheets
    well_sheet = workbook.create_sheet("well")
    lith_sheet = workbook.create_sheet("lithology")
    screen_sheet = workbook.create_sheet("screen")

    # data sheet headers added
    well_sheet.append(WELL_HEADERS)
    lith_sheet.append(LITHOLOGY_HEADERS)
    screen_sheet.append(SCREEN_HEADERS)

    for dataset in features:
        # avoid trying to process layers if they have no features.
        if not dataset.geojson:
            continue

        # set row information
        try:
            props = dataset.geojson.features[0].properties
            well_tag_number = props["well_tag_number"]

            well_values = [props.get(x, None) for x in WELL_HEADERS]
            well_sheet.append(well_values)

            lith_set = props["lithologydescription_set"]

            for item in lith_set:
                lith_values = [item.get(x, None) for x in LITHOLOGY_INDEX]
                lith_sheet.append([well_tag_number] + lith_values)

            screen_set = props["screen_set"]

            for item in screen_set:
                screen_values = [item.get(x, None) for x in SCREEN_INDEX]
                screen_sheet.append([well_tag_number] + screen_values)

        except Exception as e:
            logger.warn(e)
            continue

    # set header style for data sheets
    set_row_style(well_sheet)
    set_row_style(lith_sheet)
    set_row_style(screen_sheet)

    # fix column widths
    set_column_width(well_sheet)
    set_column_width(lith_sheet)
    set_column_width(screen_sheet)

    filename = f"{cur_date}_CrossSection"

    response = Response(
        content=save_virtual_workbook(workbook),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}.xlsx'})

    return response


def set_column_width(sheet):
    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = length


def set_row_style(sheet):
    font_header = Font(bold=True, color='FFFFFF')
    fill_header = PatternFill("solid", fgColor="4472c4")
    for row_cells in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row_cells:
            cell.font = font_header
            cell.fill = fill_header


SCREEN_INDEX = [
    'start',
    'end',
    'diameter',
    'assembly_type',
    'slot_size'
]

SCREEN_HEADERS = [
    'well_tag_number',
    'screen_from',
    'screen_to',
    'screen_diameter',
    'screen_assembly_type',
    'screen_slot_size'
]

LITHOLOGY_INDEX = [
    'start',
    'end',
    'lithology_raw_data',
    'lithology_description',
    'lithology_material',
    'lithology_hardness',
    'lithology_colour',
    'water_bearing_estimated_flow',
    'lithology_observation'
]

LITHOLOGY_HEADERS = [
    'well_tag_number',
    'lithology_from',
    'lithology_to',
    'lithology_raw_data',
    'lithology_description_code',
    'lithology_material_code',
    'lithology_hardness_code',
    'lithology_colour_code',
    'water_bearing_estimated_flow',
    'lithology_observation'
]

WELL_HEADERS = [
    "well_tag_number",
    "identification_plate_number",
    "well_identification_plate_attached",
    "well_status",
    "well_class",
    "well_subclass",
    "intended_water_use",
    "licenced_status",
    "observation_well_number",
    "obs_well_status",
    "water_supply_system_name",
    "water_supply_system_well_name",
    "street_address",
    "city",
    "legal_lot",
    "legal_plan",
    "legal_district_lot",
    "legal_block",
    "legal_section",
    "legal_township",
    "legal_range",
    "land_district",
    "legal_pid",
    "well_location_description",
    "latitude",
    "longitude",
    "utm_zone_code",
    "utm_northing",
    "utm_easting",
    "coordinate_acquisition_code",
    "construction_start_date",
    "construction_end_date",
    "alteration_start_date",
    "alteration_end_date",
    "decommission_start_date",
    "decommission_end_date",
    "diameter",
    "total_depth_drilled",
    "finished_well_depth",
    "bedrock_depth",
    "final_casing_stick_up",
    "ground_elevation",
    "ground_elevation_method",
    "static_water_level",
    "well_yield",
    "well_yield_unit",
    "artesian_flow",
    "artesian_pressure",
    "comments",
    "ems",
    "aquifer",
    "aquifer_vulnerability_index",
    "storativity",
    "transmissivity",
    "hydraulic_conductivity",
    "specific_storage",
    "specific_yield",
    "testing_method",
    "testing_duration",
    "analytic_solution_type",
    "boundary_effect",
    "aquifer_lithology",
    "well_publication_status",
]

WELL_NUMBER_COLUMNS = [
    "diameter",
    "finished_well_depth",
    "bedrock_depth",
    "ground_elevation",
    "static_water_level",
    "well_yield",
    "artesian_flow"
]

WELLS_NEARBY_HEADERS = [
    "well_tag_number",
    "latitude",
    "longitude",
    "well_yield",
    "diameter",
    "well_yield_unit",
    "finished_well_depth",
    "street_address",
    "intended_water_use",
    "aquifer_subtype",
    "aquifer_hydraulically_connected",
    "static_water_level",
    "top_of_screen",
    "top_of_screen_type",
    "distance",
    "swl_to_screen",
    "swl_to_bottom_of_well",
    "aquifer_id",
    "aquifer_material",
    "aquifer_lithology",
]


def wells_by_aquifer_xlsx_export(wells_by_aquifer: Dict):
    """
    Creates an excel file with a list of wells separated by aquifers
    Each aquifer has its on worksheet with all the corresponding wells
    Returns a response object with the excel data as content
    """

    aquifer_count = len(wells_by_aquifer)

    # There is no data
    if aquifer_count <= 0:
        return None

    workbook = openpyxl.Workbook()

    # A workbook is automatically created with 1 sheet which is set as the active one.
    # First aquifer wells goes into this sheet.
    aquifer_1, aquifer_1_wells = next(iter(wells_by_aquifer.items()))
    sheet1 = workbook.active
    sheet1.title = f"Aquifer {aquifer_1}"

    # Helper function to get row information
    def get_well_values(a_well, headers):
        well_dict = dict(a_well)

        if a_well.aquifer:
            del well_dict['aquifer']
            well_dict['aquifer_id'] = a_well.aquifer.aquifer_id
            well_dict['aquifer_material'] = a_well.aquifer.material_desc

        well_values = [well_dict.get(x, None) for x in headers]
        return well_values

    sheet1.append(WELLS_NEARBY_HEADERS)
    for well in wells_by_aquifer[aquifer_1]:
        sheet1.append(get_well_values(well, WELLS_NEARBY_HEADERS))

    sheets = {}

    # Create worksheets
    if aquifer_count > 1:
        for aquifer in list(wells_by_aquifer)[1:]:

            # Categorize unknown aquifers as 'Other'
            aquifer_sheet = f"Aquifer {aquifer}" if aquifer else 'Other'

            sheets[aquifer_sheet] = workbook.create_sheet(aquifer_sheet)
            sheets[aquifer_sheet].append(WELLS_NEARBY_HEADERS)
            for well in wells_by_aquifer[aquifer]:
                sheets[aquifer_sheet].append(get_well_values(well, WELLS_NEARBY_HEADERS))

    cur_date = datetime.datetime.now().strftime("%X-%Y-%m-%d")

    filename = f"{cur_date}_WellsNearby"
    response = Response(
        content=save_virtual_workbook(workbook),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}.xlsx'})

    return response
